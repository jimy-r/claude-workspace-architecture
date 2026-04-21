"""Bill & subscription tracker — ingests bill data, matches to services registry,
logs actuals, computes variance, emits alerts.

Inputs (bill JSON, produced by agent after Gmail extraction):
{
  "date": "2026-04-19",              # billing date
  "sender": "agl@energy.agl.com.au", # From email
  "sender_name": "AGL Energy",       # display name (optional)
  "amount": 189.40,                  # amount due
  "service_hint": "AGL Energy",      # optional; agent's best guess at which registry row
  "source_id": "gmail:msgId",        # traceability
  "subject": "Your bill"             # optional context
}

Outputs:
  1. Row appended to <project-finance>/Results/bill_actuals_log.xlsx
  2. Alert bullets appended to tasks/To Do Notes.md (## Finance & Admin), idempotent

Four alert triggers (Q9 from plan):
  (i) Actual >20% above registry `Cost`
  (ii) Sender with no matching registry row (possible new subscription)
  (iii) Renewal from a service marked Cancelled/Archived
  (iv) Duplicate bill (same sender + amount within 7 days)

Usage:
  python bill_tracker.py ensure-log
  python bill_tracker.py list-services
  python bill_tracker.py ingest bills.json
  python bill_tracker.py ingest-one --json '{...}'
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
REGISTRY_PATH = ROOT / "Reference" / "services-registry.md"
ACTUALS_LOG_PATH = ROOT / "Personal" / "Accounts" / "Results" / "bill_actuals_log.xlsx"
TODO_NOTES_PATH = ROOT / "tasks" / "To Do Notes.md"

ACTUALS_HEADERS = [
    "Date",
    "Sender",
    "Service",
    "Expected (AUD/mo)",
    "Actual",
    "Variance %",
    "Renewal",
    "Tax",
    "Notes",
    "Source ID",
]

# Rough USD→AUD conversion for normalising cost comparisons. Not precise; purpose is
# only to decide whether a bill is notably outside expected. User can revise if FX drifts.
USD_TO_AUD = 1.5
VARIANCE_THRESHOLD = 0.20  # 20% — trigger (i)
DUPLICATE_WINDOW_DAYS = 7  # trigger (iv)


@dataclass
class Service:
    """One row of the services registry."""

    name: str
    section: (
        str  # which ## section the row came from (e.g. "Personal — Utilities & Telco")
    )
    cost_raw: str
    cost_monthly_aud: float | None  # None if usage-based / free / unparseable
    billing: str
    next_renewal: str  # YYYY-MM-DD or *(TBA)* or -
    tax: str
    status: str


@dataclass
class Alert:
    """A flag for Finance & Admin review. Dedup key = (trigger, sender, date)."""

    trigger: (
        str  # "over-threshold" | "unknown-sender" | "cancelled-service" | "duplicate"
    )
    sender: str
    date_str: str
    message: str


@dataclass
class BillIngestResult:
    status: str  # "logged" | "invalid"
    bill: dict[str, Any]
    matched_service: str | None = None
    variance_pct: float | None = None
    alerts: list[Alert] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# -------------------------------- registry parser --------------------------------


def parse_registry(path: Path = REGISTRY_PATH) -> list[Service]:
    """Extract all service rows from the markdown registry."""
    text = path.read_text(encoding="utf-8")
    services: list[Service] = []
    current_section = "unknown"

    for block in _iter_sections(text):
        current_section = block["heading"]
        for row in _iter_table_rows(block["body"]):
            if len(row) < 12:
                continue
            (
                name,
                _url,
                _purpose,
                _account,
                cost_raw,
                billing,
                renewal,
                tax,
                _2fa,
                _bw,
                _rotated,
                status,
            ) = row[:12]
            if not name or name.startswith("*("):
                continue
            services.append(
                Service(
                    name=_clean(name),
                    section=current_section,
                    cost_raw=_clean(cost_raw),
                    cost_monthly_aud=_parse_cost_to_monthly_aud(cost_raw, billing),
                    billing=_clean(billing),
                    next_renewal=_clean(renewal),
                    tax=_clean(tax),
                    status=_clean(status).lower(),
                )
            )
    return services


def _iter_sections(text: str) -> list[dict[str, str]]:
    """Split markdown text into sections keyed on ## or ### headings."""
    lines = text.splitlines()
    out: list[dict[str, str]] = []
    current = {"heading": "preamble", "body": []}
    for line in lines:
        m = re.match(r"^(##+)\s+(.*)$", line)
        if m and len(m.group(1)) <= 3:
            if current["body"]:
                out.append(
                    {"heading": current["heading"], "body": "\n".join(current["body"])}
                )
            current = {"heading": m.group(2).strip(), "body": []}
        else:
            current["body"].append(line)
    if current["body"]:
        out.append({"heading": current["heading"], "body": "\n".join(current["body"])})
    return out


def _iter_table_rows(body: str) -> list[list[str]]:
    """Yield rows from any markdown table in `body`. Skips header + separator rows."""
    rows: list[list[str]] = []
    in_table = False
    for line in body.splitlines():
        if not line.strip().startswith("|"):
            in_table = False
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        # Separator row like |---|---| — skip
        if all(re.fullmatch(r":?-+:?", c) for c in cells if c):
            in_table = True
            continue
        if in_table:
            rows.append(cells)
        else:
            # First row = header, next we expect separator
            pass
    return rows


def _clean(s: str) -> str:
    return (s or "").strip()


_COST_PATTERNS = [
    # "~$200 AUD/mo", "$80 AUD/mo", "~$22 USD/mo"
    (re.compile(r"~?\$?(\d+(?:\.\d+)?)\s*(AUD|USD)?\s*/\s*mo", re.IGNORECASE), "month"),
    (re.compile(r"~?\$?(\d+(?:\.\d+)?)\s*(AUD|USD)?\s*/\s*yr", re.IGNORECASE), "year"),
    (
        re.compile(r"~?\$?(\d+(?:\.\d+)?)\s*(AUD|USD)?\s*/\s*year", re.IGNORECASE),
        "year",
    ),
]


def _parse_cost_to_monthly_aud(cost_raw: str, billing: str) -> float | None:
    """Best-effort cost parser. Returns monthly AUD; None if free/usage-based/unparseable."""
    if not cost_raw:
        return None
    low = cost_raw.lower()
    if (
        "free" in low
        or "usage" in low
        or "varies" in low
        or "per-txn" in low
        or "per-sale" in low
        or "commission" in low
    ):
        return None
    for pattern, period in _COST_PATTERNS:
        m = pattern.search(cost_raw)
        if m:
            amount = float(m.group(1))
            currency = (m.group(2) or "AUD").upper()
            if currency == "USD":
                amount *= USD_TO_AUD
            if period == "year":
                amount /= 12
            return round(amount, 2)
    # Fallback: try "$N" + billing column
    m = re.search(r"~?\$?(\d+(?:\.\d+)?)", cost_raw)
    if m:
        amt = float(m.group(1))
        if "annual" in billing.lower():
            return round(amt / 12, 2)
        if "month" in billing.lower():
            return round(amt, 2)
    return None


# ---------------------------------- matching ----------------------------------


def match_bill(bill: dict[str, Any], services: list[Service]) -> Service | None:
    """Return the best-matching Service for a bill, or None."""
    hint = (bill.get("service_hint") or "").lower()
    sender = (bill.get("sender") or "").lower()
    sender_name = (bill.get("sender_name") or "").lower()

    haystacks = [s for s in services if s.status == "live"]

    # 1. Exact or substring hint match
    if hint:
        for s in haystacks:
            if hint in s.name.lower() or s.name.lower() in hint:
                return s

    # 2. Sender-name substring match
    if sender_name:
        for s in haystacks:
            if s.name.lower() in sender_name or sender_name in s.name.lower():
                return s

    # 3. Sender local/domain substring match
    if sender and "@" in sender:
        domain = sender.split("@", 1)[1]
        parts = [domain.split(".")[0]]  # e.g. "agl"
        for s in haystacks:
            nm = s.name.lower()
            for p in parts:
                if p and (p in nm or nm.split()[0] in p):
                    return s

    return None


# ---------------------------------- log write ----------------------------------


def ensure_actuals_log(path: Path = ACTUALS_LOG_PATH) -> None:
    """Create the workbook with header row if it doesn't exist."""
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill Actuals"
    ws.append(ACTUALS_HEADERS)
    wb.save(path)


def find_duplicate(bill: dict[str, Any], log_path: Path) -> int | None:
    """Return row index of a duplicate bill in the actuals log, or None."""
    if not log_path.exists():
        return None
    target_sender = (bill.get("sender") or "").lower()
    target_amount = float(bill.get("amount") or 0)
    target_date = datetime.fromisoformat(bill["date"]).date()

    wb = load_workbook(log_path, read_only=True)
    ws = wb["Bill Actuals"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 5:
            continue
        row_date_raw, row_sender, _service, _expected, row_actual = row[:5]
        row_sender = (row_sender or "").lower()
        if row_sender != target_sender:
            continue
        try:
            row_amount = float(row_actual or 0)
        except (TypeError, ValueError):
            continue
        if abs(row_amount - target_amount) > 0.01:
            continue
        try:
            row_date = (
                row_date_raw.date()
                if hasattr(row_date_raw, "date")
                else datetime.fromisoformat(str(row_date_raw)).date()
            )
        except ValueError:
            continue
        if abs((row_date - target_date).days) <= DUPLICATE_WINDOW_DAYS:
            return row_idx
    return None


def append_actuals_row(
    bill: dict[str, Any],
    service: Service | None,
    variance_pct: float | None,
    log_path: Path,
) -> int:
    """Append a row to the actuals log. Returns the new row index."""
    ensure_actuals_log(log_path)
    wb = load_workbook(log_path)
    ws = wb["Bill Actuals"]
    row = [
        date.fromisoformat(bill["date"]),
        bill.get("sender") or bill.get("sender_name") or "",
        service.name if service else "(no match)",
        service.cost_monthly_aud
        if service and service.cost_monthly_aud is not None
        else None,
        float(bill["amount"]),
        round(variance_pct * 100, 1) if variance_pct is not None else None,
        service.next_renewal if service else "",
        service.tax if service else "",
        bill.get("subject", ""),
        bill.get("source_id", ""),
    ]
    ws.append(row)
    wb.save(log_path)
    return ws.max_row


# ---------------------------------- alerts ----------------------------------


def compute_alerts(
    bill: dict[str, Any],
    service: Service | None,
    services: list[Service],
    duplicate_row: int | None,
) -> list[Alert]:
    alerts: list[Alert] = []
    sender = bill.get("sender") or bill.get("sender_name") or "unknown"
    bdate = bill["date"]
    amount = float(bill["amount"])

    # (iv) duplicate bill
    if duplicate_row is not None:
        alerts.append(
            Alert(
                trigger="duplicate",
                sender=sender,
                date_str=bdate,
                message=f"Duplicate bill from {sender}: ${amount:.2f} on {bdate} (matches row {duplicate_row}).",
            )
        )

    # (ii) unknown sender — no match in services registry
    if service is None:
        # Check cancelled list too before declaring unknown
        cancelled = next(
            (
                s
                for s in services
                if s.status in ("cancelled", "archived")
                and s.name.lower() in sender.lower()
            ),
            None,
        )
        if cancelled:
            # (iii) cancelled-service renewal
            alerts.append(
                Alert(
                    trigger="cancelled-service",
                    sender=sender,
                    date_str=bdate,
                    message=f"Renewal from cancelled service '{cancelled.name}': ${amount:.2f} on {bdate}. Expected to be cancelled — investigate.",
                )
            )
        else:
            alerts.append(
                Alert(
                    trigger="unknown-sender",
                    sender=sender,
                    date_str=bdate,
                    message=f"Bill from sender not in services registry: {sender}, ${amount:.2f} on {bdate}. Possible new subscription.",
                )
            )
        return alerts

    # (i) over-threshold
    if service.cost_monthly_aud and service.cost_monthly_aud > 0:
        variance = (amount - service.cost_monthly_aud) / service.cost_monthly_aud
        if variance > VARIANCE_THRESHOLD:
            alerts.append(
                Alert(
                    trigger="over-threshold",
                    sender=sender,
                    date_str=bdate,
                    message=f"{service.name} bill ${amount:.2f} is {variance * 100:.0f}% over expected ~${service.cost_monthly_aud:.2f}/mo on {bdate}.",
                )
            )
    return alerts


def write_alerts_to_todo_notes(
    alerts: list[Alert], todo_path: Path = TODO_NOTES_PATH
) -> int:
    """Append unique alert bullets to `## Finance & Admin` in To Do Notes. Idempotent."""
    if not alerts:
        return 0
    if not todo_path.exists():
        raise FileNotFoundError(f"To Do Notes not found at {todo_path}")

    text = todo_path.read_text(encoding="utf-8")
    # Find the Finance & Admin section
    m = re.search(r"^## Finance & Admin\b", text, flags=re.MULTILINE)
    if not m:
        raise ValueError("'## Finance & Admin' heading not found in To Do Notes")

    # Find the insertion point — just before the next top-level heading
    after = text[m.end() :]
    next_h = re.search(r"^## ", after, flags=re.MULTILINE)
    section_end = m.end() + (next_h.start() if next_h else len(after))
    section_body = text[m.end() : section_end]

    new_bullets: list[str] = []
    for a in alerts:
        bullet = f"- [bill-monitor] {a.message}"
        if bullet in section_body:
            continue
        new_bullets.append(bullet)

    if not new_bullets:
        return 0

    insertion = "\n" + "\n".join(new_bullets) + "\n"
    new_text = text[:section_end].rstrip() + insertion + text[section_end:]
    todo_path.write_text(new_text, encoding="utf-8")
    return len(new_bullets)


# ---------------------------------- orchestration ----------------------------------


def ingest_bill(
    bill: dict[str, Any], services: list[Service], log_path: Path = ACTUALS_LOG_PATH
) -> BillIngestResult:
    errs = _validate_bill(bill)
    if errs:
        return BillIngestResult(status="invalid", bill=bill, errors=errs)

    service = match_bill(bill, services)
    dup = find_duplicate(bill, log_path)
    variance = None
    if service and service.cost_monthly_aud:
        variance = (
            float(bill["amount"]) - service.cost_monthly_aud
        ) / service.cost_monthly_aud

    alerts = compute_alerts(bill, service, services, dup)

    row = None
    if dup is None:
        row = append_actuals_row(bill, service, variance, log_path)

    return BillIngestResult(
        status="logged" if row else "duplicate",
        bill=bill,
        matched_service=service.name if service else None,
        variance_pct=variance,
        alerts=alerts,
    )


def _validate_bill(bill: dict[str, Any]) -> list[str]:
    errs: list[str] = []
    if "date" not in bill:
        errs.append("missing 'date'")
    else:
        try:
            datetime.fromisoformat(str(bill["date"]))
        except ValueError:
            errs.append(f"date not ISO: {bill['date']!r}")
    if "amount" not in bill:
        errs.append("missing 'amount'")
    else:
        try:
            if float(bill["amount"]) <= 0:
                errs.append("amount must be positive")
        except (TypeError, ValueError):
            errs.append("amount not numeric")
    if not (bill.get("sender") or bill.get("sender_name")):
        errs.append("missing 'sender' or 'sender_name'")
    if "source_id" not in bill:
        errs.append("missing 'source_id'")
    return errs


def _cli() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--registry", type=Path, default=REGISTRY_PATH)
    parser.add_argument("--log", type=Path, default=ACTUALS_LOG_PATH)
    parser.add_argument("--todo-notes", type=Path, default=TODO_NOTES_PATH)
    parser.add_argument(
        "--no-alerts", action="store_true", help="Don't write alerts to To Do Notes"
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    sub.add_parser("ensure-log", help="Create the actuals workbook if missing")
    sub.add_parser(
        "list-services", help="Parse and print the services registry as JSON"
    )

    ing = sub.add_parser("ingest", help="Ingest a JSON file of bills")
    ing.add_argument("input", type=Path)

    ione = sub.add_parser("ingest-one", help="Ingest a single bill from a JSON string")
    ione.add_argument("--json", dest="json_str", required=True)

    args = parser.parse_args()

    if args.cmd == "ensure-log":
        ensure_actuals_log(args.log)
        print(f"OK — log at {args.log}")
        return 0

    services = parse_registry(args.registry)

    if args.cmd == "list-services":
        out = [
            {
                "name": s.name,
                "section": s.section,
                "cost_monthly_aud": s.cost_monthly_aud,
                "renewal": s.next_renewal,
                "tax": s.tax,
                "status": s.status,
            }
            for s in services
        ]
        print(json.dumps(out, indent=2))
        return 0

    if args.cmd == "ingest":
        bills = json.loads(args.input.read_text(encoding="utf-8"))
    else:
        bills = [json.loads(args.json_str)]

    results: list[BillIngestResult] = [
        ingest_bill(b, services, args.log) for b in bills
    ]
    all_alerts: list[Alert] = [a for r in results for a in r.alerts]

    alerts_written = 0
    if all_alerts and not args.no_alerts:
        alerts_written = write_alerts_to_todo_notes(all_alerts, args.todo_notes)

    summary = {
        "logged": sum(1 for r in results if r.status == "logged"),
        "duplicate": sum(1 for r in results if r.status == "duplicate"),
        "invalid": sum(1 for r in results if r.status == "invalid"),
        "alerts_total": len(all_alerts),
        "alerts_written_to_notes": alerts_written,
    }

    out = [
        {
            "status": r.status,
            "matched_service": r.matched_service,
            "variance_pct": round(r.variance_pct * 100, 1)
            if r.variance_pct is not None
            else None,
            "alerts": [{"trigger": a.trigger, "message": a.message} for a in r.alerts],
            "errors": r.errors,
            "source_id": r.bill.get("source_id"),
        }
        for r in results
    ]
    print(json.dumps({"results": out, "summary": summary}, indent=2))
    return 0 if summary["invalid"] == 0 else 1


if __name__ == "__main__":
    sys.exit(_cli())
