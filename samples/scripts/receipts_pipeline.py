"""Receipt ingestion pipeline — writes receipts (email or photo) into the FY workbook.

The agent does the extraction (Gmail MCP query or Claude vision OCR) and produces a
receipts-JSON payload; this script handles:
  1. Schema validation of each receipt
  2. Categorisation via <project-finance>/Scripts/categorize.py
  3. Deduplication against existing rows in the All Transactions sheet
  4. Append + save
  5. Filing the source file (email EML/PDF or photo) to
     <project-finance>/Records/<YYYY-MM>/receipts/

Receipt schema (input JSON object):
{
  "date": "2026-04-19",           # ISO date, required
  "vendor": "Uber Eats",          # required; used for description if no description
  "amount": 24.50,                # required; always positive
  "description": "2 items",       # optional
  "account": "<Bank> Credit Card",   # optional; blank if unknown — user fills later
  "category": "Food - Delivery",  # optional; auto-derived from vendor if blank
  "source_type": "email|photo",   # required
  "source_id": "gmail:msgId|path:<file>",  # required (traceability)
  "source_path": "C:/.../file.pdf" # optional; if present, script files it
}

Usage:
  python receipts_pipeline.py ingest receipts.json
  python receipts_pipeline.py ingest-one --json '{"date":"...",...}'
  python receipts_pipeline.py validate receipts.json
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(
    0, str(Path(__file__).resolve().parent.parent / "Personal" / "Accounts" / "Scripts")
)
import categorize as _cat  # noqa: E402


WORKBOOK_PATH = (
    Path(__file__).resolve().parent.parent
    / "Personal"
    / "Accounts"
    / "Results"
    / "Financial_Summary_FY2526.xlsx"
)
RECORDS_ROOT = (
    Path(__file__).resolve().parent.parent / "Personal" / "Accounts" / "Records"
)
SHEET_NAME = "All Transactions"
REQUIRED_FIELDS = {"date", "vendor", "amount", "source_type", "source_id"}
VALID_SOURCE_TYPES = {"email", "photo"}


@dataclass
class Receipt:
    date: str
    vendor: str
    amount: float
    source_type: str
    source_id: str
    description: str = ""
    account: str = ""
    category: str = ""
    source_path: str | None = None


@dataclass
class IngestResult:
    status: str  # "appended" | "duplicate" | "invalid"
    receipt: dict[str, Any]
    row_number: int | None = None
    filed_to: str | None = None
    errors: list[str] = field(default_factory=list)


def validate_receipt(raw: dict[str, Any]) -> list[str]:
    """Return error strings; empty list = valid."""
    errors: list[str] = []
    missing = REQUIRED_FIELDS - set(raw.keys())
    if missing:
        errors.append(f"missing required fields: {sorted(missing)}")
    if "date" in raw:
        try:
            datetime.fromisoformat(str(raw["date"]))
        except ValueError:
            errors.append(f"date not ISO: {raw['date']!r}")
    if "amount" in raw:
        try:
            amt = float(raw["amount"])
            if amt <= 0:
                errors.append(f"amount must be positive: {raw['amount']}")
        except (TypeError, ValueError):
            errors.append(f"amount not numeric: {raw['amount']!r}")
    if raw.get("source_type") not in VALID_SOURCE_TYPES:
        errors.append(f"source_type must be one of {sorted(VALID_SOURCE_TYPES)}")
    return errors


def _to_receipt(raw: dict[str, Any]) -> Receipt:
    return Receipt(
        date=str(raw["date"]),
        vendor=str(raw["vendor"]).strip(),
        amount=round(float(raw["amount"]), 2),
        source_type=str(raw["source_type"]),
        source_id=str(raw["source_id"]),
        description=str(raw.get("description") or "").strip(),
        account=str(raw.get("account") or "").strip(),
        category=str(raw.get("category") or "").strip(),
        source_path=(str(raw["source_path"]) if raw.get("source_path") else None),
    )


def _build_description(receipt: Receipt) -> str:
    """Row description — vendor plus optional extra detail."""
    if receipt.description:
        return f"{receipt.vendor} — {receipt.description}"
    return receipt.vendor


def _derive_category(receipt: Receipt) -> tuple[str, bool]:
    """Return (category, exclude_from_pl). Receipts are never internal transfers,
    but categorize_transaction knows to mark some descriptions as such — honour it."""
    if receipt.category:
        return receipt.category, False
    return _cat.categorize_transaction(_build_description(receipt))


def _budget_category(category: str) -> str:
    """Coarse-grained budget bucket from a fine-grained category.

    Mirrors the conventions used in existing All Transactions rows.
    """
    if category.startswith("Food"):
        return "Food"
    if category.startswith("Transport"):
        return "Transport"
    if category.startswith("Health"):
        return "Health"
    if category.startswith("Bank"):
        return "Banking"
    if category.startswith("Income"):
        return "Income"
    if category in {
        "Mortgage",
        "Subscriptions",
        "Subscriptions - AI Tools",
        "Entertainment",
    }:
        return category.replace(" - ", " ")
    return "Other"


def _is_duplicate(ws, receipt: Receipt) -> tuple[bool, int | None]:
    """Check All Transactions for a matching (date, description, amount, account) row."""
    target_desc = _build_description(receipt).lower()
    target_date = receipt.date
    target_amount = receipt.amount
    target_account = receipt.account.lower()

    for row_idx in range(2, ws.max_row + 1):
        row_date = ws.cell(row_idx, 1).value
        row_desc = (ws.cell(row_idx, 2).value or "").lower()
        row_amount = ws.cell(row_idx, 3).value
        row_account = (ws.cell(row_idx, 5).value or "").lower()

        if hasattr(row_date, "date"):
            date_str = row_date.date().isoformat()
        elif hasattr(row_date, "isoformat"):
            date_str = row_date.isoformat()
        else:
            date_str = str(row_date or "")
        if date_str != target_date:
            continue
        if abs(float(row_amount or 0) - target_amount) > 0.01:
            continue
        if target_account and row_account and target_account != row_account:
            continue
        if row_desc == target_desc:
            return True, row_idx
    return False, None


def append_receipt(
    receipt: Receipt, workbook_path: Path = WORKBOOK_PATH
) -> IngestResult:
    """Append a single receipt to the All Transactions sheet. No-op if duplicate."""
    raw_dict = {
        "date": receipt.date,
        "vendor": receipt.vendor,
        "amount": receipt.amount,
        "source_type": receipt.source_type,
        "source_id": receipt.source_id,
    }
    wb = load_workbook(workbook_path)
    ws = wb[SHEET_NAME]

    dup, dup_row = _is_duplicate(ws, receipt)
    if dup:
        return IngestResult(status="duplicate", receipt=raw_dict, row_number=dup_row)

    category, exclude_from_pl = _derive_category(receipt)
    budget = _budget_category(category)
    new_row = [
        date.fromisoformat(receipt.date),
        _build_description(receipt),
        receipt.amount,
        "debit",
        receipt.account or "",
        category,
        exclude_from_pl,
        budget,
    ]
    ws.append(new_row)
    wb.save(workbook_path)
    return IngestResult(status="appended", receipt=raw_dict, row_number=ws.max_row)


def file_source(receipt: Receipt, records_root: Path = RECORDS_ROOT) -> Path | None:
    """Move/copy the source file to Records/<YYYY-MM>/receipts/<vendor>_<date>_<id>.<ext>.

    Returns the destination path if filed, None if no source_path provided.
    """
    if not receipt.source_path:
        return None
    src = Path(receipt.source_path)
    if not src.exists():
        raise FileNotFoundError(f"source file does not exist: {src}")

    d = date.fromisoformat(receipt.date)
    folder = records_root / f"{d.year:04d}-{d.month:02d}" / "receipts"
    folder.mkdir(parents=True, exist_ok=True)

    vendor_slug = (
        "".join(c if c.isalnum() else "-" for c in receipt.vendor).strip("-").lower()
    )
    id_slug = (
        receipt.source_id.replace(":", "_").replace("/", "_").replace("\\", "_")[:40]
    )
    dest = folder / f"{vendor_slug}_{receipt.date}_{id_slug}{src.suffix}"

    shutil.copy2(src, dest)
    return dest


def ingest_batch(
    receipts_raw: list[dict[str, Any]],
    *,
    workbook_path: Path = WORKBOOK_PATH,
    records_root: Path = RECORDS_ROOT,
    file_sources: bool = True,
) -> list[IngestResult]:
    """Validate + append + optionally file sources for a batch of receipts."""
    results: list[IngestResult] = []
    for raw in receipts_raw:
        errs = validate_receipt(raw)
        if errs:
            results.append(IngestResult(status="invalid", receipt=raw, errors=errs))
            continue
        r = _to_receipt(raw)
        result = append_receipt(r, workbook_path)
        if file_sources and result.status == "appended" and r.source_path:
            try:
                dest = file_source(r, records_root)
                result.filed_to = str(dest) if dest else None
            except Exception as e:
                result.errors.append(f"file_source failed: {e}")
        results.append(result)
    return results


def _result_to_dict(r: IngestResult) -> dict[str, Any]:
    return {
        "status": r.status,
        "source_id": r.receipt.get("source_id"),
        "vendor": r.receipt.get("vendor"),
        "amount": r.receipt.get("amount"),
        "date": r.receipt.get("date"),
        "row_number": r.row_number,
        "filed_to": r.filed_to,
        "errors": r.errors,
    }


def _cli() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--workbook", type=Path, default=WORKBOOK_PATH)
    parser.add_argument("--records-root", type=Path, default=RECORDS_ROOT)
    parser.add_argument(
        "--no-file-sources", action="store_true", help="Skip filing source files"
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    ing = sub.add_parser("ingest", help="Ingest a JSON file of receipts")
    ing.add_argument("input", type=Path, help="Path to JSON list of receipts")

    ione = sub.add_parser(
        "ingest-one", help="Ingest a single receipt from a JSON string"
    )
    ione.add_argument("--json", dest="json_str", required=True)

    val = sub.add_parser("validate", help="Validate a JSON file without ingesting")
    val.add_argument("input", type=Path)

    args = parser.parse_args()

    if args.cmd == "validate":
        items = json.loads(args.input.read_text(encoding="utf-8"))
        total_errs = 0
        for i, raw in enumerate(items):
            errs = validate_receipt(raw)
            if errs:
                total_errs += 1
                print(f"[{i}] invalid: {errs}", file=sys.stderr)
        if total_errs:
            print(f"Found {total_errs} invalid receipt(s).", file=sys.stderr)
            return 1
        print(f"OK — {len(items)} receipt(s) validate cleanly.")
        return 0

    if args.cmd == "ingest":
        items = json.loads(args.input.read_text(encoding="utf-8"))
    else:  # ingest-one
        items = [json.loads(args.json_str)]

    results = ingest_batch(
        items,
        workbook_path=args.workbook,
        records_root=args.records_root,
        file_sources=not args.no_file_sources,
    )
    print(json.dumps([_result_to_dict(r) for r in results], indent=2, default=str))
    summary = {"appended": 0, "duplicate": 0, "invalid": 0}
    for r in results:
        summary[r.status] = summary.get(r.status, 0) + 1
    print(json.dumps({"summary": summary}, indent=2), file=sys.stderr)
    return 0 if summary["invalid"] == 0 else 1


if __name__ == "__main__":
    sys.exit(_cli())
